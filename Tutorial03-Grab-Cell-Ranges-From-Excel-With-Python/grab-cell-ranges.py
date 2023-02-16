# import necessary library
# import openpyxl
# from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# Create a workbook object, since py is OOP!
# wb = Workbook()

# Load existing spreadsheet
wb = load_workbook('./Grab-Cell-Ranges-From-Excel-With-Python.xlsx')

# Create an active worksheet
## Ensure correct worksheet is activated/selected and saved in spreadsheet!
ws = wb.active

# Grab-Cell-Ranges-From-Excel-for-SingleColumn
rangeOne = ws['a1':'a7']
# print(range) # this prints Tuple(Tuple) Tuple of a Tuple. So need to loop to print its value out.

# Loop to print values from tuple of a tuple from single column
for outerTuple in rangeOne:
    for innerTuple in outerTuple:
        print(innerTuple.value)

# Grab-Cell-Ranges-From-Excel-for-TwoColumns
rangeTwo = ws['a1':'b7']

# Loop to print values from tuple of a tuple from two columns
for outerTuple in rangeTwo:
    """
    This code keeps track of the current name and balance values using separate variables.
    It then iterates over each cell in the current row and sets the name variable if the cell is in the first column,
    or the balance variable if the cell is in the second column.

    After iterating over all the cells in the row, the code checks whether both name and balance have been set to non-empty values.
    If they have, it prints out the name and balance values in the desired format.
    """
    name = ""
    balance = ""
    for innerTuple in outerTuple:
        if innerTuple.column == 1:
            name = innerTuple.value
        elif innerTuple.column == 2:
            balance = innerTuple.value
    if name != "" and balance != "":
        print(f'{name}: {balance}')



 
