# import necessary library
# import openpyxl
# from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# Create a workbook object, since py is OOP!
# wb = Workbook()

# Load existing spreadsheet
wb = load_workbook('./How-to-Load-Excel-in-Python-Openpyxl-Tutorial-1.xlsx')

# Create an active worksheet
ws = wb.active

# Set a variable name for column header in the spreadsheet
## Remember to set all formula into values only in the spreadsheet, otherwise only the formula will be printed.

name = ws["a2"].value
balance = ws["b2"].value

# Print sth frm our spreadsheet
## if PermissionError: [Errno 13] Permission denied: 'fileName.xlsx', simply close the spreadsheet and rerun the program.
# print(ws['a2']) # this print its object
# print(ws['a2'].value) # this print its object's value
# print(ws['a2'].value) # this print its object's value
# print(f'{}:{}') # use f-string to print more column
# print(f'{ws["a2"].value}:{ws["b2"].value}') # use f-string to print more column example w/o variable name
print(f'{name}:{balance}') # use f-string to print more column example w variable name


# wb is define by convention to refer to excel or any spreadsheet workbook 
# filepath: can be absolute or relative filepath.
# Tips: wb in same directory can use relative path ./filename or simply 'filename'
